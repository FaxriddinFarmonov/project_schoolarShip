# views.py
import json
from datetime import date, datetime
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from app.models.read_terminal import TerminalRead


def export_terminal_to_excel_flat_address(request, pk):
    try:
        terminal = TerminalRead.objects.get(id=pk)
    except TerminalRead.DoesNotExist:
        return HttpResponse("Bunday ID topilmadi!", status=404)

    addr = terminal.address or {}
    addr_keys = list(addr.keys())

    headers = [
        "ID", "Terminal ID", "Terminal Name", "Class GUID", "Inst Name",
        "Status", "Title", "Notes", "Create Time", "Create Day", "Activate Time",
        "Default CCY", "Default Language"
    ]
    headers += [f"Address:{k}" for k in addr_keys]
    headers += [
        "Branch ID", "Branch RID", "Branch Code", "Owner RID",
        "Last RQ Time", "Last Resp Time", "Last Online RRN",
        "MAC Error Cnt", "Host Password", "Issuer Cards",
        "Term Enhanced Password", "Created At"
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)

    def safe(v):
        """Excel uchun xavfsiz qiymat"""
        if v is None:
            return ""
        if isinstance(v, (dict, list)):
            return json.dumps(v, ensure_ascii=False)
        if isinstance(v, (datetime, date)):
            return v.strftime("%Y-%m-%d %H:%M:%S")
        return v

    row = [
        safe(terminal.id),
        safe(terminal.terminal_id),
        safe(terminal.terminal_name),
        safe(terminal.class_guid),
        safe(terminal.inst_name),
        safe(terminal.status),
        safe(terminal.title),
        safe(terminal.notes),
        safe(terminal.create_time),
        safe(terminal.create_day),
        safe(terminal.activate_time),
        safe(terminal.default_ccy),
        safe(terminal.default_language),
    ]
    for k in addr_keys:
        row.append(safe(addr.get(k, "")))

    row += [
        safe(terminal.branch_id),
        safe(terminal.branch_rid),
        safe(terminal.branch_code),
        safe(terminal.owner_rid),
        safe(terminal.last_rq_time),
        safe(terminal.last_resp_time),
        safe(terminal.last_online_rrn),
        safe(terminal.mac_error_cnt),
        safe(terminal.host_password),
        safe(terminal.issuer_cards),
        safe(terminal.term_enhanced_password),
        safe(terminal.created_at),
    ]

    ws.append(row)

    # ustun kengliklarini moslash
    for i, column_cells in enumerate(ws.columns, 1):
        max_length = max((len(str(cell.value)) if cell.value is not None else 0) for cell in column_cells)
        ws.column_dimensions[get_column_letter(i)].width = min(50, max_length + 2)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename=terminal_{pk}_flat.xlsx'
    wb.save(response)
    return response
